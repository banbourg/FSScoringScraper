#!/bin/env python
# coding: utf-8

import os
import glob
import re
import pandas as pd
from openpyxl import load_workbook
import numpy as np

# TO DOs: (1) Fix missing deductions

def return_isu_abbrev(s):
    temp = filter(None, re.split(r'(\d+)', s))
    return temp[0]

def is_nan(x):
    return (x is np.nan or x != x)

def return_row_list(i, k_min, df, list):
    for k in range(k_min, len(df.columns)):
        if df.iloc[i, k] is not None and not is_nan(df.iloc[i, k]):
            list.append(df.iloc[i, k])
    return list

def clean_elt_name(cur_string, replace_list):
    for cur_word in replace_list:
        cur_string = cur_string.replace(cur_word, '')
    return cur_string

def add_segment_identifiers(df, identifiers, segment_competitors_list):
    df['discipline'] = identifiers[0]
    df.set_index('discipline', append=True, inplace=True)
    df['category'] = identifiers[1]
    df.set_index('category', append=True, inplace=True)
    df['season'] = identifiers[2]
    df.set_index('season', append=True, inplace=True)
    df['event'] = identifiers[3]
    df.set_index('event', append=True, inplace=True)
    df['team_event'] = identifiers[4]
    df.set_index('team_event', append=True, inplace=True)
    df['skater'] = segment_competitors_list[-1][3]
    df.set_index('skater', append=True, inplace=True)
    df['segment'] = identifiers[5]
    df.set_index('segment', append=True, inplace=True)

def main():
    read_path = os.path.expanduser('~/Desktop/bias/pdftoxls/')
    write_path = os.path.expanduser('~/Desktop/bias/output/')

    event_dic = {'gpjpn': 'NHK', 'gpfra': 'TDF', 'gpcan': 'SC', 'gprus': 'COR', 'gpusa': 'SA', 'gpchn': 'COC',
                 'gpf': 'GPF', 'wc': 'WC', 'fc': '4CC', 'owg': 'OWG', 'wtt': 'WTT'}
    calls = ['!', 'e', '<', '<<', '*', '+REP', 'V', 'x', 'X']

    year_regex = re.compile(r'\d+')
    combo_regex = re.compile(r'\+[0-9]')

    files = sorted(glob.glob(read_path + '*.xlsx'))

    all_scores_list = []
    all_pcs_list = []
    all_goe_list = []
    all_calls_list = []
    all_deductions_list = []
    all_competitors_list = []

    for f in files:
        filename = f[43:]  # 43

        # 1. DERIVE YEAR
        year_list = [int(x) for x in year_regex.findall(filename)]
        if len(year_list) > 1:
            print 'Error - MULTIPLE YEARS LISTED IN FILENAME ', filename
            exit()
        else:
            year_data = year_list[0]
            if len(str(year_data)) == 2:
                event_year = 2000 + year_data
            elif len(str(year_data)) == 4 and str(year_data)[:2] == '20':
                event_year = year_data
            elif len(str(year_data)) == 4 and int(str(year_data)[:2]) == (int(str(year_data)[-2:]) - 1):
                event_year = 2000 + int(str(year_data)[:2])
            else:
                print 'Error - SOMETHING WONKY WITH DATE FORMATTING IN FILENAME ', filename
                exit()

        # 2. DERIVE EVENT & SUB EVENT
        event = event_dic[return_isu_abbrev(filename.lower())]
        team_event = 'Team' if 'Team' in filename else ''

        # 3. DERIVE SEASON
        if event in ['4CC', 'OWG', 'WC', 'WTT']:
            season = "SB" + str(event_year - 1)
        else:
            season = "SB" + str(event_year)

        # 4. TAG SEGMENT
        segment = 'SP' if 'SP' in filename else 'FS'

        # 5. TAG DISCIPLINE AND CATEGORY
        discipline = 'Men' if 'Men' in filename else 'Ladies'
        category = 'Jr' if 'Junior' in filename else 'Sr'
        dc_short = category + discipline[0]

        identifiers = [discipline, category, season, event, team_event, segment]

        print 'SUMMARY: ', filename, season, event, team_event, event_year, discipline, category, segment

        wb = load_workbook(f)
        segment_competitors_list = []
        # segment_goe_list = []
        # segment_calls_list = []
        # segment_pcs_list = []
        # segment_scores_list = []
        segment_deductions_list = []
        segment_exploded_names = []

        for sheet in wb.sheetnames:
            # print sheet
            ws = wb[sheet]

            length = {'JrLSP': 7, 'JrMSP': 7, 'SrLSP': 7, 'SrMSP': 7, 'JrLFS': 11, 'JrMFS': 12, 'SrLFS': 12,
                      'SrMFS': 13}

            raw_df = pd.DataFrame(ws.values)
            for i in raw_df.index:
                for j in raw_df.columns:

                    # SCRAPE COMPETITOR NAME
                    if raw_df.iloc[i, j] == 'Name':
                        name_row = []
                        for k in range(i + 2, i + 5):
                            start = max(j-2, 0)
                            for l in range(start, j + 2):
                                namelike_regex = re.search(r'[A-Z]{2,}', unicode(raw_df.iloc[k, l]))
                                if namelike_regex is not None:
                                    return_row_list(k, 0, raw_df, name_row)
                                    break
                            if name_row:
                                break
                        assert name_row

                        # The 'fuck your names, Dutch people' exception - they break the pdf conversion
                        spaced_patronym_regex = re.search(r'^\d+\s+\D+', unicode(name_row[0]))
                        if spaced_patronym_regex is not None:
                            e_handler = unicode(name_row[0]).split(' ', 1)
                            name_row[0] = int(e_handler[0])
                            name_row.insert(1, e_handler[1])
                        # Check name order
                        exploded_name = name_row[1].split(' ')
                        first_name_list = [word for word in exploded_name if unicode(word[1]).islower()]
                        first_name = ' '.join(first_name_list)
                        last_name_list = [word for word in exploded_name if unicode(word[1]).isupper()]
                        last_name = ' '.join(last_name_list)
                        short_last_name = ''.join(last_name_list)

                        competitor_name = first_name + ' ' + last_name

                        country = 'RUS' if name_row[2] == 'OAR' else name_row[2]

                        segment_competitors_list.append((season, discipline, category, competitor_name, country))
                        segment_exploded_names.append((first_name, short_last_name))

                    # SCRAPE PCS SCORES
                    elif 'Skating Skills' in unicode(raw_df.iloc[i, j]):
                        single_pcs_list = []
                        for k in range(i, i + 5):
                            raw_row_data = []
                            return_row_list(k, j + 1, raw_df, raw_row_data)
                            row_data = []
                            for raw_cell in raw_row_data:
                                cleaner = [u.replace(u',',u'.') for u in str(raw_cell).split()]
                                for v in cleaner:
                                    try:
                                        cleanest_cell = float(v)
                                    except:
                                        cleanest_cell = ''
                                row_data.append(cleanest_cell)
                            row_data = filter(None, row_data)

                            single_pcs_list.append(row_data[1:-1])

                        single_pcs_df = pd.DataFrame(single_pcs_list, index=['ss', 'tr', 'pc', 'ch', 'in'],
                                                     columns=['j1', 'j2', 'j3', 'j4', 'j5', 'j6', 'j7', 'j8', 'j9'])
                        single_pcs_df.rename_axis('judge', axis='columns', inplace=True)
                        single_pcs_df.rename_axis('component', axis='index', inplace=True)

                        add_segment_identifiers(single_pcs_df, identifiers, segment_competitors_list)
                        segment_pcs_list.extend([single_pcs_df])

                    # SCRAPE DEDUCTIONS
                    # For clarity, formatting issues we're trying to tackle:
                    #    Falls may or may not be followed (or preceded) by # of falls in parentheses
                    #    Total fall deduction may not equal # of falls * -1 (can add deductions for interruption)
                    #    Some rows have totals, some don't
                    elif 'Deductions' in unicode(raw_df.iloc[i, j]) and j < 4:
                        ded_row = []
                        return_row_list(i, j, raw_df, ded_row)
                        # Stringify and remove number of falls in brackets, split
                        ded_row = [re.sub(r'\(\d+\)', '', str(ded)) for ded in ded_row]
                        ded_list = []
                        for ded in ded_row:
                            ded_list.extend(re.split('[,!: ]+', str(ded)))
                        ded_list = filter(None, ded_list[1:-1]) # Gets rid of initial 'deduction' heading & total
                        # Do this, given the existence of strings: ded_list = filter(lambda x: abs(float(x)) < 10, ded_list)
                        for x in ded_list:
                            try:
                                if abs(float(x)) > 15:
                                    ded_list.remove(x)
                            except:
                                pass

                        i2 = 0
                        while i2 < len(ded_list): # can I do p directly instead of indexing e.g p+1?
                            # Ensure all number are negative
                            digits1 = re.search(r'[\d+]',ded_list[i2])
                            if digits1 is not None:
                                # Falls: -2.00(2)
                                ded_list[i2] = -1*float(ded_list[i2]) if float(ded_list[i2]) > 0 else float(ded_list[i2])

                            if (i2+1) < (len(ded_list)-1):
                                digits2 = re.search(r'[\d+]',ded_list[i2+1])
                                if digits1 is None and digits2 is None and ded_list[i2+1] != 'Total':
                                    ded_list = [' '.join(ded_list[i2:i2+2])] + ded_list[i2+2:]
                            i2 += 1
                        ded_tuples = zip(ded_list[0::2],ded_list[1::2])

                        for (ded_type, ded_points) in ded_tuples:
                            segment_deductions_list.append((discipline, category, season, event, team_event,
                                                            segment_competitors_list[-1][3], segment) +
                                                           (ded_type, ded_points))

                    # SCRAPE ELEMENTS, CALLS, GOE AND TES SCORES
                    elif 'Elements' in unicode(raw_df.iloc[i, j]):
                        single_goe_list = []
                        single_calls_list = []
                        single_scores_list = []
                        elt_id_list = []
                        competitor_short_name = segment_exploded_names[-1][1] + segment_exploded_names[-1][0][0]

                        # Identify whether elt list starts on next line or not
                        incr = 1 if raw_df.iloc[i + 1, j] is not None else 2
                        for k in range(i + incr, i + incr + length[dc_short + segment]):
                            elt_row = []
                            return_row_list(k, 0, raw_df, elt_row)

                            if len(elt_row) >= 13:  # Sometimes people forget an element apparently
                                elt_no = int(elt_row[0])

                                # Get rid of random columns of dashes
                                if elt_row[-2] == '-' and elt_row[-3] != '-':
                                    del elt_row[-2]

                                # Clean elt name and separate Sq and level
                                elt_name = clean_elt_name(elt_row[1], calls)
                                lvl_regex = re.search(r'\d+$', elt_row[1])
                                if lvl_regex is not None:
                                    level = lvl_regex.group(0)
                                    elt_name = elt_name[:-1]
                                else:
                                    level = ''

                                # POPULATE TECH CALL FLAGS
                                invalid = 1 if any('*' in str(cell) for cell in elt_row) else 0
                                h2 = 1 if any('x' in str(cell) for cell in elt_row) else 0

                                if len(combo_regex.findall(elt_name)) > 0 or '+COMBO' in elt_name:
                                    combo_flag = 1
                                else:
                                    combo_flag = 0

                                # Note: Distinction between UR and Downgrade was brought in from SB2011
                                # multiple calls per jumping pass
                                if int(season[-2:]) < 11 and any('<' in str(cell) for cell in elt_row):
                                    downgrade_flag = 1
                                elif any('<<' in str(cell) for cell in elt_row):
                                    downgrade_flag = 1
                                else:
                                    downgrade_flag = 0

                                severe_edge_flag = 1 if any('e' in str(cell) for cell in elt_row) else 0
                                unclear_edge_flag = 1 if any('!' in str(cell) for cell in elt_row) else 0
                                rep_flag = 1 if any('+REP' in str(cell) for cell in elt_row) else 0
                                failed_spin = 1 if any('V' in str(cell) for cell in elt_row) else 0

                                if combo_flag == 1:
                                    jumps = elt_row[1].split('+')

                                    jump_1 = clean_elt_name(jumps[0], calls)
                                    jump_2 = clean_elt_name(jumps[1], calls)
                                    jump_3 = clean_elt_name(jumps[2], calls) if len(jumps) > 2 else ''
                                    jump_4 = clean_elt_name(jumps[3], calls) if len(jumps) > 3 else ''

                                    j1_sev_edge = 1 if 'e' in jumps[0] else 0
                                    j2_sev_edge = 1 if 'e' in jumps[1] else 0
                                    j3_sev_edge = 1 if (len(jumps) > 2 and 'e' in jumps[2]) else 0
                                    j4_sev_edge = 1 if (len(jumps) > 3 and 'e' in jumps[3]) else 0

                                    j1_unc_edge = 1 if '!' in jumps[0] else 0
                                    j2_unc_edge = 1 if '!' in jumps[1] else 0
                                    j3_unc_edge = 1 if (len(jumps) > 2 and '!' in jumps[2]) else 0
                                    j4_unc_edge = 1 if (len(jumps) > 3 and '!' in jumps[3]) else 0

                                    j1_down = 1 if '<<' in jumps[0] else 0
                                    j2_down = 1 if '<<' in jumps[1] else 0
                                    j3_down = 1 if (len(jumps) > 2 and '<<' in jumps[2]) else 0
                                    j4_down = 1 if (len(jumps) > 3 and '<<!' in jumps[3]) else 0

                                    j1_ur = 1 if ('<' in jumps[0] and '<<' not in jumps[0]) else 0
                                    j2_ur = 1 if ('<' in jumps[1] and '<<' not in jumps[1]) else 0
                                    j3_ur = 1 if (len(jumps) > 2 and '<' in jumps[2] and '<<' not in jumps[2]) else 0
                                    j4_ur = 1 if (len(jumps) > 3 and '<' in jumps[3] and '<<' not in jumps[3]) else 0

                                else:
                                    jump_1 = clean_elt_name(elt_row[1], calls) if lvl_regex is None else ''
                                    j1_sev_edge = severe_edge_flag
                                    j1_unc_edge = unclear_edge_flag
                                    j1_ur = 1 if downgrade_flag == 0 and any(
                                        '<' in str(cell) for cell in elt_row) else 0
                                    j1_down = downgrade_flag
                                    jump_2, j2_sev_edge, j2_unc_edge, j2_ur = '', 0, 0, 0
                                    j2_down, jump_3, j3_sev_edge, j3_unc_edge = 0, '', 0, 0
                                    j3_ur, j3_down, jump_4, j4_sev_edge = 0, 0, '', 0
                                    j4_unc_edge, j4_ur, j4_down = 0, 0, 0

                                ur_flag = 1 if 1 in [j1_ur, j2_ur, j3_ur, j4_ur] else 0

                                temp_numbers = []
                                for cell in elt_row[2:-10]:
                                    temp_numbers.extend(unicode(cell).split(' '))
                                numbers = [unicode(cell).strip() for cell in temp_numbers if cell not in calls]
                                for call_notation in calls:
                                    numbers[0] = numbers[0].replace(call_notation, '').strip()
                                elt_bv = float(numbers[0])
                                elt_sov_goe, elt_total = float(numbers[1]), float(elt_row[-1])

                                # SCRAPE GOE SCORES
                                goe_row = []
                                for b in elt_row[-10:-1]:
                                    try:
                                        clean = int(b)
                                    except:
                                        clean = 0
                                    goe_row.append(clean)
                                # print 'goe_row: ', goe_row

                            else:
                                elt_no = k - i
                                elt_name = 'MISSING_ELEMENT'
                                level, h2, combo_flag, ur_flag = 0, 0, 0, 0
                                downgrade_flag, severe_edge_flag, unclear_edge_flag = 0, 0, 0
                                rep_flag, called_jumps, invalid, failed_spin = 0, 0, 0, 0

                                jump_1, j1_sev_edge, j1_unc_edge, j1_ur = '', 0, 0, 0
                                j1_down, jump_2, j2_sev_edge, j2_unc_edge = 0, '', 0, 0
                                j2_ur, j2_down, jump_3, j3_sev_edge = 0, 0, '', 0
                                j3_unc_edge, j3_ur, j3_down, jump_4 = 0, 0, 0, ''
                                j4_sev_edge, j4_unc_edge, j4_ur, j4_down = 0, 0, 0, 0
                                elt_bv, elt_sov_goe, elt_total = 0, 0, 0
                                goe_row = [0, 0, 0, 0, 0, 0, 0, 0, 0]

                            elt_id = 'SB' + season[-2:] + event + team_event[:1].upper() + dc_short \
                                     + competitor_short_name + segment + str(elt_no)
                            elt_id_list.append(elt_id)
                            # print elt_id

                            calls_row = (elt_no, elt_name, level, invalid, h2, combo_flag, ur_flag, downgrade_flag,
                                         severe_edge_flag, unclear_edge_flag, rep_flag, jump_1, j1_sev_edge,
                                         j1_unc_edge,
                                         j1_ur, j1_down, jump_2, j2_sev_edge, j2_unc_edge, j2_ur, j2_down, jump_3,
                                         j3_sev_edge, j3_unc_edge, j3_ur, j3_down, jump_4,
                                         j4_sev_edge, j4_unc_edge, j4_ur, j4_down, failed_spin)
                            scores_row = (elt_name, level, h2, elt_bv, elt_sov_goe, elt_total)

                            single_scores_list.append(scores_row)
                            single_calls_list.append(calls_row)
                            single_goe_list.append(goe_row)

                        call_cols = ['elt_no', 'elt_name', 'level', 'invalid', 'h2', 'combo_flag', 'ur_flag',
                                     'downgrade_flag', 'severe_edge_flag', 'unclear_edge_flag', 'rep_flag', 'jump_1',
                                     'j1_sev_edge', 'j1_unc_edge', 'j1_ur', 'j1_down', 'jump_2', 'j2_sev_edge',
                                     'j2_unc_edge', 'j2_ur', 'j2_down', 'jump_3', 'j3_sev_edge', 'j3_unc_edge', 'j3_ur',
                                     'j3_down', 'jump_4', 'j4_sev_edge', 'j4_unc_edge', 'j4_ur','j4_down', 'failed_spin']
                        single_calls_df = pd.DataFrame(single_calls_list, index=elt_id_list, columns=call_cols)
                        # print single_calls_df

                        single_scores_df = pd.DataFrame(single_scores_list, index=elt_id_list,
                                                        columns=['elt_name', 'level', 'h2', 'elt_bv', 'elt_sov_goe',
                                                                 'elt_total'])
                        # print single_scores_df

                        single_goe_df = pd.DataFrame(single_goe_list, index=elt_id_list, columns=['j1', 'j2', 'j3', 'j4', 'j5', 'j6',
                                                                                                  'j7', 'j8', 'j9'])
                        single_goe_df.rename_axis('judge', axis='columns', inplace=True)
                        # print single_goe_df

                        # ADD THE OTHER INFO COLUMNS - Figure how to loop through the dfs without python thinking
                        add_segment_identifiers(single_scores_df, identifiers, segment_competitors_list)
                        add_segment_identifiers(single_goe_df, identifiers, segment_competitors_list)
                        add_segment_identifiers(single_calls_df, identifiers, segment_competitors_list)

                        segment_scores_list.append(single_scores_df)
                        segment_goe_list.append(single_goe_df)
                        segment_calls_list.append(single_calls_df)

        # print type(segment_scores_list.stack())
        segment_competitors_df = pd.DataFrame(segment_competitors_list,
                                              columns=['season', 'disc', 'category', 'name', 'country'])

        segment_deductions_df = pd.DataFrame(segment_deductions_list,
                                             columns=['discipline', 'category', 'season', 'event', 'team_event',
                                                      'skater', 'segment', 'ded_type', 'ded_points'])

        segment_scores_df = pd.concat(segment_scores_list)  # .stack()
        segment_pcs_df = pd.concat(segment_pcs_list).stack()
        segment_pcs_df.name = 'pcs'
        segment_goe_df = pd.concat(segment_goe_list).stack()
        segment_goe_df.name = 'goe'
        segment_calls_df = pd.concat(segment_calls_list)  # .stack()

        all_scores_list.append(segment_scores_df)
        all_pcs_list.append(segment_pcs_df)
        all_goe_list.append(segment_goe_df)
        all_calls_list.append(segment_calls_df)
        all_deductions_list.append(segment_deductions_df)
        all_competitors_list.append(segment_competitors_df)
        print '        loaded full segment df into overall summary list'

    all_scores_df = pd.concat(all_scores_list)
    print 'scores df concatenated'
    all_pcs_df = pd.concat(all_pcs_list)
    all_pcs_df = all_pcs_df.reset_index()
    print 'pcs df concatenated'
    all_goe_df = pd.concat(all_goe_list)
    all_goe_df = all_goe_df.reset_index()
    print 'goe df concatenated'
    all_calls_df = pd.concat(all_calls_list)
    print 'calls df concatenated'
    all_deductions_df = pd.concat(all_deductions_list)
    all_deductions_df = all_deductions_df.reset_index(drop=True)
    print 'deductions df concatenated'
    all_competitors_df = pd.concat(all_competitors_list)
    all_competitors_df.drop_duplicates(subset=['category', 'name', 'country'], keep='last', inplace=True)
    all_competitors_df = all_competitors_df.reset_index(drop=True)
    print 'competitors df concatenated'



    date = '180622'
    ver = '1'
    all_scores_df.to_csv(write_path + 'scores_'+date+ver+'.csv', mode='a', encoding='utf-8', header=True)
    all_pcs_df.to_csv(write_path + 'pcs_'+date+ver+'.csv', mode='a', encoding='utf-8', header=True)
    all_goe_df.to_csv(write_path + 'goe_'+date+ver+'.csv', mode='a', encoding='utf-8',header=True)
    all_calls_df.to_csv(write_path + 'calls_'+date+ver+'.csv', mode='a', encoding='utf-8', header=True)
    all_deductions_df.to_csv(write_path + 'deductions_'+date+ver+'.csv', mode='a', encoding='utf-8', header=True)
    all_competitors_df.to_csv(write_path + 'competitors_'+date+ver+'.csv', mode='a', encoding='utf-8', header=True)


main()
